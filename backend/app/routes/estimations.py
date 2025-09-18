from __future__ import annotations

from datetime import datetime
from typing import List

from bson import ObjectId
from fastapi import APIRouter, Depends, HTTPException, status, UploadFile, File
from fastapi.responses import FileResponse, JSONResponse

from app.core.security import get_current_user_id
from app.db.mongo import get_db
from app.models.estimation import Estimation, EstimationVersion, Feature, ResourceAllocation, ReviewRecord
from app.services.estimations import (
    add_review,
    create_estimation,
    get_estimation,
    list_estimations,
    list_versions,
    rollback_version,
    snapshot_version,
    update_features,
    update_resources,
    update_estimation_title_client_desc,
    delete_estimation,
    update_envelope_data,
)
from app.services.excel import ExcelService
from io import BytesIO
from openpyxl import load_workbook
import hashlib
import asyncio


router = APIRouter()


async def get_current_user_role_dep(user_id: str = Depends(get_current_user_id)) -> str:
    db = get_db()
    doc = await db.users.find_one({"_id": ObjectId(user_id)}, {"role": 1})
    if not doc:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="User not found")
    return str(doc.get("role", "")).lower()


async def _list_estims(user_id: str, role: str) -> List[Estimation]:
    # Estimators see only their own; ops/admin see all
    if role == "estimator":
        return await list_estimations(created_by=user_id)
    return await list_estimations()


@router.get("/", response_model=List[Estimation])
async def get_all(user_id: str = Depends(get_current_user_id), role: str = Depends(get_current_user_role_dep)) -> List[Estimation]:
    return await _list_estims(user_id, role)


# Removed duplicate GET without slash to reduce redundancy


@router.post("/", response_model=Estimation)
async def create(payload: Estimation, user_id: str = Depends(get_current_user_id), role: str = Depends(get_current_user_role_dep)) -> Estimation:
    
    now = datetime.utcnow()
    payload.created_at = now
    payload.updated_at = now
    payload.creator_id = user_id
    if payload.current_version is None:
        payload.current_version = EstimationVersion(
            version_number=1,
            features=[],
            resources=[],
            created_by=user_id,
            created_at=now,
        )
    if payload.versions is None:
        payload.versions = []
    if payload.review_records is None:
        payload.review_records = []
    return await create_estimation(payload)


@router.get("/{estimation_id}", response_model=Estimation)
async def get_one(estimation_id: str) -> Estimation:
    est = await get_estimation(estimation_id)
    if est is None:
        raise HTTPException(status_code=404, detail="Not found")
    return est


@router.patch("/{estimation_id}", response_model=Estimation)
async def update_basic(
    estimation_id: str,
    payload: dict,
    role: str = Depends(get_current_user_role_dep)
) -> Estimation:
    # Allow admin, estimator, ops to update; but only admin can change creator_id
    if role not in ("estimator", "ops", "admin"):
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Not allowed")
    if "creator_id" in payload and role != "admin":
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Only admin can change estimator")
    est = await update_estimation_title_client_desc(estimation_id, payload)
    if est is None:
        raise HTTPException(status_code=404, detail="Not found")
    return est


@router.delete("/{estimation_id}")
async def remove(estimation_id: str, role: str = Depends(get_current_user_role_dep)) -> dict:
    if role not in ("estimator", "ops", "admin"):
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Not allowed to delete")
    ok = await delete_estimation(estimation_id)
    if not ok:
        raise HTTPException(status_code=404, detail="Not found")
    return {"deleted": True}


@router.put("/{estimation_id}/features", response_model=Estimation)
async def set_features(estimation_id: str, features: List[Feature], role: str = Depends(get_current_user_role_dep)) -> Estimation:
    if role not in ("estimator", "ops", "admin"):
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Only estimator, ops or admin can modify features")
    est = await update_features(estimation_id, features)
    if est is None:
        raise HTTPException(status_code=404, detail="Not found")
    return est


@router.put("/{estimation_id}/resources", response_model=Estimation)
async def set_resources(estimation_id: str, resources: List[ResourceAllocation], role: str = Depends(get_current_user_role_dep)) -> Estimation:
    if role not in ("estimator", "ops", "admin"):
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Not allowed to modify resources")
    est = await update_resources(estimation_id, resources)
    if est is None:
        raise HTTPException(status_code=404, detail="Not found")
    return est


# Lightweight create-or-update by envelope_data payload
@router.post("/import-envelope")
async def import_envelope(payload: dict, user_id: str = Depends(get_current_user_id)) -> dict:
    """Create a new estimation from an envelope-style JSON (project + rows).
    Returns the created estimation id.
    """
    from app.services.importer import build_estimation_from_envelope
    from pymongo.errors import DuplicateKeyError
    from app.db.mongo import get_db
    from datetime import datetime

    est = build_estimation_from_envelope(payload, creator_id=user_id)
    db = get_db()

    # Try to find an existing temporary estimation with the same title
    existing_temp = await db.estimations.find_one({
        "title": est.title,
        "is_temporary": True
    })

    estimation_id_to_return = None
    if existing_temp:
        # Overwrite the existing temporary estimation
        existing_id = existing_temp["_id"]
        update_doc = est.model_dump(by_alias=True)
        if "_id" in update_doc:
            del update_doc["_id"]  # Don't try to update the _id
        
        await db.estimations.update_one({"_id": existing_id}, {"$set": update_doc})
        estimation_id_to_return = str(existing_id)
    else:
        # No existing temporary one, so try to create a new one
        try:
            created = await create_estimation(est)
            estimation_id_to_return = created.id or getattr(created, "_id", None)
        except DuplicateKeyError:
            # A finalized estimation with this title exists. Append suffix and retry.
            est.title = f"{est.title} - Copy {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')}"
            created = await create_estimation(est)
            estimation_id_to_return = created.id or getattr(created, "_id", None)

    if estimation_id_to_return:
        est_obj = await get_estimation(estimation_id_to_return)
        if est_obj:
            asyncio.create_task(ExcelService.generate_excel(est_obj))

    return {"estimation_id": estimation_id_to_return, "temporary": True}


@router.post("/{estimation_id}/finalize")
async def finalize_estimation(estimation_id: str, user_id: str = Depends(get_current_user_id)) -> dict:
    """Mark a temporary estimation as finalized so it becomes visible in listings."""
    db = get_db()
    from bson import ObjectId
    await db.estimations.update_one({"_id": ObjectId(estimation_id)}, {"$set": {"is_temporary": False}})
    return {"ok": True}


@router.get("/{estimation_id}/export-excel")
async def export_excel(estimation_id: str) -> FileResponse:
    est = await get_estimation(estimation_id)
    if est is None:
        raise HTTPException(status_code=404, detail="Not found")
    path = await ExcelService.generate_excel(est)
    if not path:
        raise HTTPException(status_code=400, detail="Failed to generate Excel")
    return FileResponse(path, filename=f"{est.title}_FILLED_{estimation_id}.xlsx")


@router.post("/{estimation_id}/upload-excel")
async def upload_excel(estimation_id: str, file: UploadFile = File(...)) -> dict:
    """Accept an uploaded Excel and return a minimal mapping summary.
    For MVP, parse a sheet to extract resources table if present.
    """
    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="Empty file")

    # Load current estimation to build mapping reference
    est = await get_estimation(estimation_id)
    if est is None:
        raise HTTPException(status_code=404, detail="Not found")
    existing_rows = (est.envelope_data.rows if est.envelope_data else []) or []

    def _hash_key(platform: str, module: str, component: str, feature: str) -> str:
        raw = "|".join([str(platform or "").strip().lower(), str(module or "").strip().lower(), str(component or "").strip().lower(), str(feature or "").strip().lower()])
        return hashlib.sha1(raw.encode("utf-8")).hexdigest()

    # Build lookup by row_id or by hash
    id_to_index: dict[str, int] = {}
    hash_to_index: dict[str, int] = {}
    for idx, r in enumerate(existing_rows):
        rid = getattr(r, "row_id", None)
        if rid:
            id_to_index[str(rid)] = idx
        key = _hash_key(getattr(r, "platform", ""), getattr(r, "module", ""), getattr(r, "component", ""), getattr(r, "feature", ""))
        hash_to_index[key] = idx

    # Parse workbook
    wb = load_workbook(filename=BytesIO(content), data_only=True)
    ws = wb.active
    headers = [str(c.value).strip() if c.value is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]

    # Expected minimal headers matching sample: adjust names as needed
    # We look for these canonical names
    def _col(name: str) -> int:
        try:
            return headers.index(name)
        except ValueError:
            return -1

    col_row_id = _col("Row ID")
    col_platform = _col("Platform (Desktop / Web / Mobile)")
    if col_platform < 0:
        col_platform = _col("Platform")
    col_module = _col("Module")
    col_component = _col("Component")
    col_feature = _col("Features") if _col("Features") >= 0 else _col("Feature")
    col_make = _col("Make/ Reuse") if _col("Make/ Reuse") >= 0 else _col("Make/Reuse")
    col_complexity = _col("Complexity (Simple / Complex / Average)")
    if col_complexity < 0:
        col_complexity = _col("Complexity")

    matched = 0
    updated = 0
    unmatched = 0
    updated_rows: list[dict] = []

    for row in ws.iter_rows(min_row=2):
        vals = [c.value for c in row]
        platform = str(vals[col_platform]).strip() if col_platform >= 0 and vals[col_platform] is not None else ""
        module = str(vals[col_module]).strip() if col_module >= 0 and vals[col_module] is not None else ""
        component = str(vals[col_component]).strip() if col_component >= 0 and vals[col_component] is not None else ""
        feature = str(vals[col_feature]).strip() if col_feature >= 0 and vals[col_feature] is not None else ""
        make_reuse = str(vals[col_make]).strip() if col_make >= 0 and vals[col_make] is not None else ""
        complexity = str(vals[col_complexity]).strip() if col_complexity >= 0 and vals[col_complexity] is not None else ""
        rid = str(vals[col_row_id]).strip() if col_row_id >= 0 and vals[col_row_id] is not None else None

        target_index = None
        if rid and rid in id_to_index:
            target_index = id_to_index[rid]
        else:
            key = _hash_key(platform, module, component, feature)
            target_index = hash_to_index.get(key)

        if target_index is None:
            unmatched += 1
            continue

        matched += 1
        # Record an updated minimal row payload (only six fields + row_id)
        new_row = {
            "row_id": rid or _hash_key(platform, module, component, feature),
            "platform": platform,
            "module": module,
            "component": component,
            "feature": feature,
            "make_or_reuse": make_reuse if make_reuse in ("Make", "Reuse") else ("Make" if make_reuse.lower().startswith("m") else "Reuse" if make_reuse.lower().startswith("r") else "Make"),
            "complexity": complexity if complexity in ("Simple", "Average", "Complex") else "Average",
        }
        updated_rows.append(new_row)
        updated += 1

    # Parse global Resources table from any sheet
    parsed_resources: list[dict] = []
    def _find_resources_in_sheet(sheet) -> list[dict]:
        max_scan_rows = min(100, sheet.max_row)
        for r in range(1, max_scan_rows + 1):
            row_vals = [str(c.value).strip().lower() if c.value is not None else "" for c in sheet[r]]
            if not row_vals:
                continue
            if ("resources" in row_vals) and ("days" in row_vals) and ("no. of resources" in row_vals or "number of resources" in row_vals or "# of resources" in row_vals or "no of resources" in row_vals) and ("allocation" in row_vals):
                # Build header index map
                hdr_idx = {v: i for i, v in enumerate(row_vals)}
                def _idx(*names: str) -> int | None:
                    for n in names:
                        if n in hdr_idx:
                            return hdr_idx[n]
                    return None
                i_resource = _idx("resources")
                i_days = _idx("days")
                i_count = _idx("no. of resources", "number of resources", "# of resources", "no of resources")
                i_alloc = _idx("allocation")
                out: list[dict] = []
                for rr in range(r + 1, sheet.max_row + 1):
                    cells = sheet[rr]
                    get = lambda idx: (cells[idx].value if (idx is not None and idx < len(cells)) else None)
                    name = get(i_resource)
                    if name is None or str(name).strip() == "":
                        # stop at first empty row in resources column
                        break
                    days_v = get(i_days)
                    cnt_v = get(i_count)
                    alloc_v = str(get(i_alloc) or "").strip().lower()
                    allocation_type = "pt" if alloc_v.startswith("part") else ("ft" if alloc_v.startswith("full") else "pt")
                    try:
                        days_i = int(float(days_v or 0))
                    except Exception:
                        days_i = 0
                    try:
                        cnt_i = int(float(cnt_v or 0))
                    except Exception:
                        cnt_i = 0
                    out.append({"role": str(name), "days": days_i, "count": cnt_i, "allocation_type": allocation_type})
                return out
        return []

    # Search all sheets for resources table
    for sheet in wb.worksheets:
        found = _find_resources_in_sheet(sheet)
        if found:
            parsed_resources = found
            break

    return {"matched": matched, "updated": updated, "unmatched": unmatched, "rows": updated_rows, "resources": parsed_resources}


@router.post("/{estimation_id}/populate-from-excel")
async def populate_from_excel(estimation_id: str, payload: dict) -> JSONResponse:
    """Finalize populate: update resources and optionally envelope rows from mapped data.
    Only updates provided fields; preserves others.
    """
    updates: dict = {}
    if "resources" in payload:
        try:
            from app.models.estimation import ResourceAllocation
            resources = [ResourceAllocation(**r) for r in payload["resources"]]
            est = await update_resources(estimation_id, resources)
            if est is None:
                raise HTTPException(status_code=404, detail="Not found")
            updates["resources"] = len(resources)
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Invalid resources: {e}")
    if "rows" in payload:
        try:
            est = await get_estimation(estimation_id)
            if est is None:
                raise HTTPException(status_code=404, detail="Not found")
            env = est.envelope_data.model_dump() if est.envelope_data else {"schema_version": "1.0", "project": {}, "rows": []}
            env["rows"] = payload["rows"]
            est2 = await update_envelope_data(estimation_id, env)
            if est2 is None:
                raise HTTPException(status_code=404, detail="Not found")
            updates["rows"] = len(payload["rows"])
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Invalid rows: {e}")
    return JSONResponse({"ok": True, **updates})
# Update full envelope_data for an estimation
@router.put("/{estimation_id}/envelope", response_model=Estimation)
async def set_envelope(
    estimation_id: str,
    payload: dict,
    user_id: str = Depends(get_current_user_id),
    role: str = Depends(get_current_user_role_dep),
) -> Estimation:
    # Check permissions: Admin and Ops can edit any. Estimators can only edit their own.
    if role not in ("admin", "ops"):
        if role == "estimator":
            existing_est = await get_estimation(estimation_id)
            if not existing_est:
                raise HTTPException(status_code=404, detail="Not found")
            if existing_est.creator_id != user_id:
                raise HTTPException(
                    status_code=status.HTTP_403_FORBIDDEN,
                    detail="Estimators can only modify their own estimations.",
                )
        else:
            # Deny any other role
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="You do not have permission to modify the envelope.",
            )

    est = await update_envelope_data(estimation_id, payload)
    if est is None:
        raise HTTPException(status_code=404, detail="Not found")
    return est


# Removed reviewer-only review endpoint as 'reviewer' role is no longer used


@router.post("/{estimation_id}/versions", response_model=Estimation)
async def create_version(estimation_id: str, user_id: str = Depends(get_current_user_id), role: str = Depends(get_current_user_role_dep)) -> Estimation:
    if role not in ("estimator", "ops", "admin"):
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Only estimator, ops or admin can snapshot versions")
    est = await snapshot_version(estimation_id, user_id)
    if est is None:
        raise HTTPException(status_code=404, detail="Not found")
    return est


@router.get("/{estimation_id}/versions", response_model=List[EstimationVersion])
async def versions(estimation_id: str) -> List[EstimationVersion]:
    return await list_versions(estimation_id)


@router.post("/{estimation_id}/versions/{version_number}/rollback", response_model=Estimation)
async def rollback(estimation_id: str, version_number: int, role: str = Depends(get_current_user_role_dep)) -> Estimation:
    if role not in ("estimator", "ops", "admin"):
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Only estimator, ops or admin can rollback versions")
    est = await rollback_version(estimation_id, version_number)
    if est is None:
        raise HTTPException(status_code=404, detail="Not found")
    return est


