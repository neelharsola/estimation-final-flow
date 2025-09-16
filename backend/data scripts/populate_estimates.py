from __future__ import annotations
import argparse
import json
import re
from pathlib import Path
from typing import Dict, Any, List, Optional, Union
from datetime import datetime

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

TARGET_SHEET = "Estimation"

def norm(s: str) -> str:
    s = s.strip().lower()
    s = s.replace("\n", " ")
    s = re.sub(r"[\u2013\u2014\-–—]+", "-", s)
    s = re.sub(r"[^a-z0-9#/\.\-\s()]+", "", s)
    s = re.sub(r"\s+", " ", s)
    return s

HEADER_TO_JSON = {
    norm("Platform (Desktop / Web / Mobile)"): "platform",
    norm("Module"): "module",
    norm("Component"): "component",
    norm("Features"): "feature",
    norm("Make/ Reuse"): "make_or_reuse",
    norm("Complexity (Simple / Complex / Average)"): "complexity",
    norm("Project Name"): "previous_project_actual.project_name",
    norm("Actual (working day)"): "previous_project_actual.actual_working_days",
    norm("UI Design"): "hours.ui_design",
    norm("UI Module"): "hours.ui_module",
    norm("BL"): "hours.backend_logic",
    norm("General"): "hours.general",
    norm("Service/ API"): "hours.service_api",
    norm("DB Struct."): "hours.db_structure",
    norm("DB Prog."): "hours.db_programming",
    norm("DB - UDF"): "hours.db_udf",
    norm("# Comp."): "num_components",
}

JSON_NUMERIC_KEYS = {
    "hours.ui_design", "hours.ui_module", "hours.backend_logic", "hours.general",
    "hours.service_api", "hours.db_structure", "hours.db_programming", "hours.db_udf",
    "previous_project_actual.actual_working_days", "num_components"
}

def safe_excel_value(value: Any) -> Union[str, int, float, bool, None]:
    if value is None:
        return None
    
    if isinstance(value, (int, float)):
        if isinstance(value, float):
            if value != value or value == float('inf') or value == float('-inf'):
                return None
            if not (-1.7976931348623157e+308 <= value <= 1.7976931348623157e+308):
                return str(value)
        return value
    
    if isinstance(value, bool):
        return value
    
    if isinstance(value, str):
        value = value.replace('\x00', '')
        value = re.sub(r'[\x01-\x08\x0B\x0C\x0E-\x1F\x7F-\x9F]', '', value)
        
        if len(value) > 32767:
            value = value[:32767]
        
        return value
    
    if isinstance(value, list):
        try:
            if all(isinstance(x, str) for x in value):
                joined = " | ".join(str(x) for x in value)
                return safe_excel_value(joined)
            else:
                json_str = json.dumps(value, ensure_ascii=False)
                return safe_excel_value(json_str)
        except Exception:
            return str(value)
    
    if isinstance(value, dict):
        try:
            json_str = json.dumps(value, ensure_ascii=False)
            return safe_excel_value(json_str)
        except Exception:
            return str(value)
    
    try:
        result = str(value)
        return safe_excel_value(result)
    except Exception:
        return ""

def safe_numeric_convert(value: Any) -> Optional[float]:
    if value is None or value == "":
        return None
    
    if isinstance(value, (int, float)):
        if isinstance(value, float) and (value != value or value == float('inf') or value == float('-inf')):
            return None
        return float(value)
    
    if isinstance(value, str):
        value = value.strip()
        if value == "":
            return None
        try:
            result = float(value)
            if result != result or result == float('inf') or result == float('-inf'):
                return None
            return result
        except (ValueError, OverflowError):
            return None
    
    return None

def flatten_row(row: Dict[str, Any]) -> Dict[str, Any]:
    """Flatten nested JSON structure into flat dictionary"""
    out: Dict[str, Any] = {}
    
    simple_fields = ["platform", "module", "component", "feature", "make_or_reuse", "complexity", "num_components"]
    for k in simple_fields:
        if k in row:
            out[k] = row[k]

    ppa = row.get("previous_project_actual")
    if isinstance(ppa, dict):
        out["previous_project_actual.project_name"] = ppa.get("project_name")
        out["previous_project_actual.actual_working_days"] = ppa.get("actual_working_days")
    else:
        out["previous_project_actual.project_name"] = None
        out["previous_project_actual.actual_working_days"] = None

    hours = row.get("hours", {}) or {}
    hour_fields = ["ui_design", "ui_module", "backend_logic", "general", 
                   "service_api", "db_structure", "db_programming", "db_udf"]
    
    for hk in hour_fields:
        value = hours.get(hk, 0) if isinstance(hours, dict) else 0
        out[f"hours.{hk}"] = value
    
    return out

def scan_header_columns(ws: Worksheet, search_rows: int = 20) -> Dict[str, int]:
    header_map: Dict[str, int] = {}
    
    try:
        max_col = min(ws.max_column or 1, 200)
        max_search_row = min(search_rows, ws.max_row or 1)
        
        for r in range(1, max_search_row + 1):
            for c in range(1, max_col + 1):
                try:
                    cell = ws.cell(row=r, column=c)
                    val = cell.value
                    if val is None:
                        continue
                        
                    key = norm(str(val))
                    if key and key not in header_map:
                        header_map[key] = c
                except Exception:
                    continue
                    
    except Exception as e:
        print(f"Warning: Error scanning headers: {e}")
        
    return header_map

def locate_columns(ws: Worksheet) -> Dict[str, int]:
    scan = scan_header_columns(ws)
    col_map: Dict[str, int] = {}
    
    for header_norm, json_key in HEADER_TO_JSON.items():
        if header_norm in scan:
            col_map[json_key] = scan[header_norm]
    
    return col_map

def find_header_row(ws: Worksheet) -> int:
    targets = {norm("module"), norm("features"), norm("platform (desktop / web / mobile)")}
    max_search_row = min(30, ws.max_row or 1)
    
    for r in range(1, max_search_row + 1):
        seen = set()
        try:
            max_col_to_check = min(ws.max_column or 1, 50)
            for c in range(1, max_col_to_check + 1):
                cell = ws.cell(row=r, column=c)
                val = cell.value
                if val is not None:
                    seen.add(norm(str(val)))
        except Exception:
            continue
            
        if targets.issubset(seen):
            return r
    
    for r in range(1, min(10, ws.max_row + 1)):
        try:
            max_col_to_check = min(ws.max_column or 1, 20)
            if any(ws.cell(row=r, column=c).value not in (None, "") 
                   for c in range(1, max_col_to_check + 1)):
                return r
        except Exception:
            continue
    
    return 1

def next_empty_data_row(ws: Worksheet, data_cols: List[int], header_row: int) -> int:
    start = header_row + 1
    max_check_row = min((ws.max_row or 0) + 100, start + 5000)
    
    for r in range(start, max_check_row):
        try:
            is_empty = True
            for c in data_cols:
                if c > 1000:
                    continue
                cell_value = ws.cell(row=r, column=c).value
                if cell_value not in (None, ""):
                    is_empty = False
                    break
            if is_empty:
                return r
        except Exception:
            continue
    
    return (ws.max_row or 0) + 1

def get_last_filled_data_row(ws: Worksheet, data_cols: List[int], header_row: int) -> int:
    last = header_row
    max_check_row = min((ws.max_row or 0) + 1, header_row + 5000)
    
    for r in range(header_row + 1, max_check_row):
        try:
            has_data = False
            for c in data_cols:
                if c > 1000:
                    continue
                cell_value = ws.cell(row=r, column=c).value
                if cell_value not in (None, ""):
                    has_data = True
                    break
            if has_data:
                last = r
        except Exception:
            continue
    
    return last

def copy_formulas(ws: Worksheet, src_row: int, dst_row: int):
    try:
        max_col_to_copy = min(ws.max_column + 1, 100)
        
        for c in range(1, max_col_to_copy):
            try:
                src_cell = ws.cell(row=src_row, column=c)
                dst_cell = ws.cell(row=dst_row, column=c)
                
                if (dst_cell.value in (None, "") and 
                    src_cell.value is not None and 
                    isinstance(src_cell.value, str) and 
                    src_cell.value.startswith("=")):
                    dst_cell.value = src_cell.value
                    
            except Exception as e:
                print(f"Warning: Could not copy formula from ({src_row}, {c}) to ({dst_row}, {c}): {e}")
                continue
                
    except Exception as e:
        print(f"Warning: Error in copy_formulas: {e}")

def write_rows_to_estimation(ws: Worksheet, rows: List[Dict[str, Any]]):
    if not rows:
        return
    
    try:
        header_row = find_header_row(ws)
        col_map = locate_columns(ws)
        
        owned_keys = [
            "platform", "module", "component", "feature", "make_or_reuse", "complexity",
            "previous_project_actual.project_name", "previous_project_actual.actual_working_days",
            "hours.ui_design", "hours.ui_module", "hours.backend_logic", "hours.general",
            "hours.service_api", "hours.db_structure", "hours.db_programming", "hours.db_udf",
            "num_components"
        ]
        
        data_cols = [col_map[k] for k in owned_keys if k in col_map]
        
        if not data_cols:
            raise RuntimeError(f"Could not locate required columns in '{TARGET_SHEET}'. Found columns: {list(col_map.keys())}")

        start_row = next_empty_data_row(ws, data_cols, header_row)
        last_filled = get_last_filled_data_row(ws, data_cols, header_row)

        current_row = start_row
        
        for i, row in enumerate(rows):
            try:
                flat = flatten_row(row)
                
                if current_row > ws.max_row:
                    ws.cell(row=current_row, column=1)
                
                if current_row > last_filled and last_filled > header_row:
                    copy_formulas(ws, last_filled, current_row)

                for json_key, col_idx in col_map.items():
                    if col_idx > 1000:
                        continue
                        
                    val = flat.get(json_key)
                    if val is None:
                        continue
                    
                    if json_key in JSON_NUMERIC_KEYS:
                        numeric_val = safe_numeric_convert(val)
                        if numeric_val is not None:
                            val = numeric_val
                    
                    safe_val = safe_excel_value(val)
                    
                    try:
                        ws.cell(row=current_row, column=col_idx, value=safe_val)
                    except Exception as e:
                        print(f"Warning: Could not write value to cell ({current_row}, {col_idx}): {e}")
                        try:
                            ws.cell(row=current_row, column=col_idx, value=str(safe_val) if safe_val is not None else "")
                        except Exception:
                            continue
                
                current_row += 1
                
            except Exception as e:
                print(f"Warning: Error processing row {i + 1}: {e}")
                continue
                
    except Exception as e:
        raise RuntimeError(f"Error writing rows to worksheet: {e}")

def main():
    parser = argparse.ArgumentParser(description="Populate 'Estimation' sheet without overwriting, keeping formulas intact.")
    parser.add_argument("--json", required=True, help="Path to JSON envelope with 'rows'.")
    parser.add_argument("--inbook", required=True, help="Path to input workbook (.xlsx or .xlsm).")
    parser.add_argument("--outbook", required=False, help="Path to output workbook. Default: .FILLED before extension.")
    args = parser.parse_args()

    inbook = Path(args.inbook)
    if not inbook.exists():
        raise FileNotFoundError(f"Input workbook not found: {inbook}")

    if args.outbook:
        outbook = Path(args.outbook)
    else:
        timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
        outbook = inbook.with_name(inbook.stem + f".FILLED-{timestamp}" + inbook.suffix)

    try:
        with open(args.json, "r", encoding="utf-8") as f:
            data = json.load(f)
    except json.JSONDecodeError as e:
        raise ValueError(f"Invalid JSON format: {e}")
    except Exception as e:
        raise ValueError(f"Could not read JSON file: {e}")

    if not isinstance(data, dict) or "rows" not in data:
        raise ValueError("JSON must contain a 'rows' key")
    
    if not isinstance(data["rows"], list):
        raise ValueError("JSON 'rows' must be an array")

    try:
        wb = openpyxl.load_workbook(str(inbook), data_only=False)
    except Exception as e:
        try:
            wb = openpyxl.load_workbook(str(inbook), data_only=False, keep_vba=True)
        except Exception as e2:
            raise ValueError(f"Could not load workbook {inbook}. Error: {e}")

    if TARGET_SHEET not in wb.sheetnames:
        raise RuntimeError(f"Sheet '{TARGET_SHEET}' not found. Available sheets: {wb.sheetnames}")

    ws = wb[TARGET_SHEET]
    
    try:
        write_rows_to_estimation(ws, data["rows"])
    except Exception as e:
        raise RuntimeError(f"Error writing data to worksheet: {e}")

    try:
        wb.save(str(outbook))
    except PermissionError:
        raise PermissionError(f"Cannot write to {outbook}. File may be open in Excel or you lack permissions.")
    except Exception as e:
        raise RuntimeError(f"Could not save workbook to {outbook}: {e}")

    print(f"Successfully wrote {len(data['rows'])} row(s) to '{TARGET_SHEET}' sheet.")
    print(f"Output workbook: {outbook}")

if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(f"Error: {e}")
        exit(1)